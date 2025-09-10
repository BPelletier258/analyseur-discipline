# === APP ANALYSEUR DISCIPLINE – Filtrage par article ==================================
# Version : 2025‑09‑08 (stabilisée)
# Points clés :
#  - Case à cocher TOUJOURS visible (isole les segments des 4 colonnes d'intérêt)
#  - Listes à puces conservées (plus de "\n" affichés) pour les colonnes textuelles
#  - Aucune puce pour les colonnes à valeur unique (noms, numéros, dates, totaux…)
#  - Mise en évidence de l'article recherché (rouge) dans 5 colonnes :
#       • Liste des chefs et articles en infraction
#       • Nbr Chefs par articles
#       • Nbr Chefs par articles par période de radiation
#       • Nombre de chefs par articles et total amendes
#       • Nombre de chefs par article ayant une réprimande
#  - Remplacement systématique de NaN par « — »
#  - Format des montants dans « Total amendes » : « 500 $ », « 5 000 $ », etc.
#  - En‑tête figée dans la vue (position: sticky)
#  - Export Excel :
#       • 1re ligne : « Article filtré : <valeur> »
#       • En‑tête figée (freeze panes sur la 2ᵉ ligne)
#       • Largeurs auto‑ajustées

import io
import os
import re
import math
import time
import unicodedata
from datetime import datetime
from typing import Dict, Optional, Set, List, Tuple

import pandas as pd
from flask import Flask, request, render_template_string, send_file

app = Flask(__name__)

# ──────────────────────────────────────────────────────────────────────────────
#  Styles
# ──────────────────────────────────────────────────────────────────────────────
STYLE = r"""
<style>
:root{
  --w-def: 22rem;           /* largeur par défaut */
  --w-2x: 44rem;            /* colonnes élargies x2 */
  --w-num: 8rem;            /* colonnes numériques étroites */
}
body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;margin:22px}
h1{font-size:20px;margin:0 0 10px}
.note{background:#fff8e0;border:1px solid #ffd48a;padding:8px 10px;border-radius:6px;margin:10px 0}
form{display:grid;gap:10px;margin:0 0 12px}
input[type="text"], input[type="file"]{font-size:14px}
.kbd{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;background:#f3f4f6;padding:2px 4px;border-radius:4px}
.download{margin:10px 0}

/* Tableau */
.table-wrap{height:60vh;overflow:auto;border:1px solid #ddd}
.table{border-collapse:collapse;width:100%;table-layout:fixed}
th,td{border:1px solid #ddd;padding:6px 8px;vertical-align:top}
th{background:#f3f4f6;position:sticky;top:0;z-index:1;text-align:center}

/* Largeurs ciblées */
.col-def{min-width:var(--w-def)}
.col-2x{min-width:var(--w-2x)}
.col-num{min-width:var(--w-num);text-align:center}

/* Listes à puces homogènes */
.bullets{margin:0;padding-left:1.1rem}
.bullets li{margin:0.15rem 0}

/* Mise en évidence de l'article */
.hl{color:#d00;font-weight:600}

/* Messages */
.msg{white-space:pre-wrap;font-family:ui-monospace,Menlo,monospace;font-size:12px}
.ok{color:#065f46}.err{color:#7f1d1d}
</style>
"""

HTML = r"""
<!doctype html>
<html><head><meta charset="utf-8"><title>Analyseur Discipline – Filtrage par article</title>
{{ style|safe }}</head>
<body>
  <h1>Analyseur Discipline – Filtrage par article</h1>
  <div class="note">
    Règles : détection exacte de l’article. Si la 1<sup>re</sup> cellule contient « <span class="kbd">Article filtré :</span> », on ignore la 1<sup>re</sup> ligne (lignes d’en‑tête sur la 2<sup>e</sup>).
  </div>
  <form method="POST" enctype="multipart/form-data">
    <label>Article à rechercher (ex. <span class="kbd">29</span>, <span class="kbd">59(2)</span>)</label>
    <input name="article" type="text" value="{{ article or '' }}" required>

    <label><input type="checkbox" name="isolate" {% if isolate %}checked{% endif %}> Afficher uniquement le segment contenant l’article dans les 4 colonnes d’intérêt</label>

    <label>Fichier Excel</label>
    <input name="file" type="file" accept=".xlsx,.xlsm" required>

    <button>Analyser</button>
    <div>Formats : .xlsx / .xlsm</div>
  </form>

  {% if table_html %}
    <div class="download"><a href="{{ download_url }}">Télécharger le résultat (Excel)</a></div>
    <div class="table-wrap">{{ table_html|safe }}</div>
  {% endif %}

  {% if message %}
    <p class="msg {{ 'ok' if ok else 'err' }}">{{ message }}</p>
  {% endif %}
</body></html>
"""

# ──────────────────────────────────────────────────────────────────────────────
#  Normalisation & résolveur d’en‑têtes
# ──────────────────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    if not isinstance(s, str):
        s = '' if s is None else str(s)
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = s.replace('\u00A0',' ')
    return ' '.join(s.strip().lower().split())

# Alias de colonnes (souples)
ALIASES: Dict[str, Set[str]] = {
    # 5 colonnes où l’on surligne l’article
    'liste_chefs_articles': {_norm('Liste des chefs et articles en infraction')},
    'nbr_chefs_articles': {_norm('Nbr Chefs par articles'), _norm('Nombre de chefs par articles')},
    'nbr_chefs_par_radiation': {_norm('Nbr Chefs par articles par période de radiation')},
    'nb_chefs_et_total_amendes': {_norm('Nombre de chefs par articles et total amendes')},
    'nb_chefs_ayant_reprimande': {_norm('Nombre de chefs par article ayant une réprimande')},

    # Autres colonnes textuelles souvent en listes
    'resume_faits': {_norm('Résumé des faits concis')},
    'liste_sanctions': {_norm('Liste des sanctions imposées')},
    'autres_mesures': {_norm('Autres mesures ordonnées')},
    'a_verifier': {_norm('À vérifier'), _norm('A verifier')},

    # Colonnes à valeur unique (jamais en puces)
    'nom_intime': {_norm("Nom de l'intimé"), _norm("Nom de l’intimé")},
    'ordre_pro': {_norm('Ordre professionnel')},
    'numero_decision': {_norm('Numéro de la décision'), _norm('Numero de la decision')},
    'date_rendue': {_norm('Date de la décision rendue')},
    'nature_decision': {_norm('Nature de la décision')},
    'periode_faits': {_norm('Période des faits')},
    'plaidoyer': {_norm('Plaidoyer de culpabilité')},
    'total_chefs': {_norm('Total chefs')},
    'radiation_max': {_norm('Radiation max')},
    'total_amendes': {_norm('Total amendes')},
    'total_reprimandes': {_norm('Total réprimandes')},
    'date_creation': {_norm('Date de création')},
    'date_maj': {_norm('Date de mise à jour')},
}

# 4 colonnes d’intérêt pour le filtrage/isolation
FILTER_KEYS = [
    'nbr_chefs_articles',
    'nbr_chefs_par_radiation',
    'nb_chefs_et_total_amendes',
    'nb_chefs_ayant_reprimande',
]

# Colonnes jamais en puces
NO_BULLETS_KEYS = {
    'nom_intime','ordre_pro','numero_decision','date_rendue','nature_decision','periode_faits',
    'plaidoyer','total_chefs','radiation_max','total_amendes','total_reprimandes','date_creation','date_maj'
}

# Colonnes pour surlignage de l’article
HILITE_KEYS = {
    'liste_chefs_articles','nbr_chefs_articles','nbr_chefs_par_radiation',
    'nb_chefs_et_total_amendes','nb_chefs_ayant_reprimande'
}

# Mapping de classes de largeur (HTML)
COL_WIDTH_CLASS = {
    'resume_faits':'col-2x',
    'autres_mesures':'col-2x',
    'numero_decision':'col-def',
    'date_creation':'col-def',
    'date_maj':'col-def',
    'nbr_chefs_articles':'col-2x',
    'nbr_chefs_par_radiation':'col-2x',
    'nb_chefs_et_total_amendes':'col-2x',
    'nb_chefs_ayant_reprimande':'col-2x',
    'total_chefs':'col-num','radiation_max':'col-num','total_amendes':'col-num','total_reprimandes':'col-num'
}

# Libellés canoniques → affichage
DISPLAY_LABEL = {
    'nom_intime': "Nom de l'intimé",
    'ordre_pro': 'Ordre professionnel',
    'numero_decision': 'Numéro de la décision',
    'date_rendue': 'Date de la décision rendue',
    'nature_decision': 'Nature de la décision',
    'periode_faits': 'Période des faits',
    'plaidoyer': 'Plaidoyer de culpabilité',
    'resume_faits': 'Résumé des faits concis',
    'liste_chefs_articles': 'Liste des chefs et articles en infraction',
    'nbr_chefs_articles': 'Nbr Chefs par articles',
    'total_chefs': 'Total chefs',
    'liste_sanctions': 'Liste des sanctions imposées',
    'nbr_chefs_par_radiation': 'Nbr Chefs par articles par période de radiation',
    'radiation_max': 'Radiation max',
    'nb_chefs_et_total_amendes': 'Nombre de chefs par articles et total amendes',
    'total_amendes': 'Total amendes',
    'nb_chefs_ayant_reprimande': 'Nombre de chefs par article ayant une réprimande',
    'total_reprimandes': 'Total réprimandes',
    'autres_mesures': 'Autres mesures ordonnées',
    'a_verifier': 'À vérifier',
    'date_creation': 'Date de création',
    'date_maj': 'Date de mise à jour',
}

ALL_KEYS_ORDER = [
    'nom_intime','ordre_pro','numero_decision','date_rendue','nature_decision','periode_faits','plaidoyer',
    'resume_faits','liste_chefs_articles','nbr_chefs_articles','total_chefs','liste_sanctions',
    'nbr_chefs_par_radiation','radiation_max','nb_chefs_et_total_amendes','total_amendes',
    'nb_chefs_ayant_reprimande','total_reprimandes','autres_mesures','a_verifier','date_creation','date_maj'
]

# ──────────────────────────────────────────────────────────────────────────────
#  Lecture Excel (règle « Article filtré : »)
# ──────────────────────────────────────────────────────────────────────────────

def read_excel(file_stream) -> pd.DataFrame:
    preview = pd.read_excel(file_stream, header=None, nrows=1, engine='openpyxl')
    file_stream.seek(0)
    if not preview.empty and isinstance(preview.iloc[0,0], str) and _norm(preview.iloc[0,0]).startswith(_norm('Article filtré :')):
        df = pd.read_excel(file_stream, header=0, skiprows=1, engine='openpyxl')
    else:
        df = pd.read_excel(file_stream, header=0, engine='openpyxl')
    return df

# ──────────────────────────────────────────────────────────────────────────────
#  Résolution d’en‑têtes
# ──────────────────────────────────────────────────────────────────────────────

def resolve_headers(df: pd.DataFrame) -> Dict[str,str]:
    norm2orig = {_norm(c): c for c in df.columns}
    out: Dict[str,str] = {}
    for key, variants in ALIASES.items():
        for v in variants:
            if v in norm2orig:
                out[key] = norm2orig[v]
                break
    return out

# ──────────────────────────────────────────────────────────────────────────────
#  Recherche article & helpers
# ──────────────────────────────────────────────────────────────────────────────

def build_article_regex(art: str) -> re.Pattern:
    token = (art or '').strip()
    if not token:
        raise ValueError('Article vide')
    esc = re.escape(token)
    tail = r"(?![\d.])" if token[-1].isdigit() else r"\b"
    return re.compile(rf"(?:\b(?:art(?:icle)?\s*[: ]*)?)({esc}){tail}", re.I)

# Nettoyage minimal pour la recherche

def _prep(v: object) -> str:
    if not isinstance(v, str):
        v = '' if pd.isna(v) else str(v)
    return (
        v.replace('•','\n').replace('·','\n').replace('◦','\n')
         .replace('\u00A0',' ').replace('\u202F',' ')
         .replace('\r\n','\n').replace('\r','\n')
    )

# Segmentation en items texte

def split_items(text: str) -> List[str]:
    text = _prep(text)
    parts = re.split(r"\n+|;", text)
    return [p.strip() for p in parts if p and p.strip()]

# Extraction des seuls segments contenant l’article

def keep_segments_with_article(text: str, rx: re.Pattern) -> List[str]:
    return [seg for seg in split_items(text) if rx.search(seg)]

# Mise en évidence HTML

def hilite_html(text: str, rx: re.Pattern) -> str:
    def repl(m: re.Match) -> str:
        return f"<span class=\"hl\">{m.group(1)}</span>"
    return rx.sub(repl, text)

# Format montant « 12 500 $ »

def format_money(val) -> str:
    if pd.isna(val):
        return '—'
    try:
        n = float(val)
    except Exception:
        # non numérique → on renvoie tel quel
        return str(val)
    if math.isnan(n):
        return '—'
    s = f"{int(round(n)):,.0f}".replace(',', ' ')
    return f"{s} $"

# ──────────────────────────────────────────────────────────────────────────────
#  Construction du tableau d’affichage & d’export
# ──────────────────────────────────────────────────────────────────────────────

def build_tables(df: pd.DataFrame, colmap: Dict[str,str], art: str, isolate: bool) -> Tuple[pd.DataFrame, pd.DataFrame, str]:
    rx = build_article_regex(art)

    # Filtrage des lignes : au moins une des 4 colonnes contient l’article
    masks = []
    for key in FILTER_KEYS:
        col = colmap.get(key)
        if col and col in df.columns:
            masks.append(df[col].astype(str).apply(lambda v: bool(rx.search(_prep(v)))))
    if not masks:
        raise ValueError("Aucune des 4 colonnes d’intérêt n’a été trouvée dans le fichier.")
    mask_any = masks[0]
    for m in masks[1:]:
        mask_any = mask_any | m
    df = df[mask_any].copy()
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), "Aucune ligne ne contient l’article dans les colonnes cibles."

    # Prépare DataFrame d’export (texte brut) et d’affichage (HTML)
    excel_rows: List[Dict[str,str]] = []
    html_rows: List[Dict[str,str]] = []

    for _, row in df.iterrows():
        excel_row: Dict[str,str] = {}
        html_row: Dict[str,str] = {}

        for key in ALL_KEYS_ORDER:
            label = DISPLAY_LABEL[key]
            src = colmap.get(key)
            val = '' if not src else row.get(src, '')
            val_str = '' if pd.isna(val) else str(val)

            # Colonnes des montants totaux
            if key == 'total_amendes':
                excel_val = format_money(val)
                html_val = excel_val
            else:
                # Colonnes listes/texte vs simples
                if key in NO_BULLETS_KEYS:
                    excel_val = '—' if (not val_str.strip()) else val_str.strip()
                    # surlignage si concerné (ex. aucun de ces simples n’est dans HILITE_KEYS)
                    html_val = excel_val
                else:
                    # Liste : soit on isole (4 colonnes), soit on garde tout
                    items: List[str]
                    if isolate and key in FILTER_KEYS:
                        items = keep_segments_with_article(val_str, rx)
                    else:
                        items = split_items(val_str)
                    if not items:
                        excel_val = '—'
                        html_val = '—'
                    else:
                        # Surlignage si la colonne est dans la liste
                        if key in HILITE_KEYS:
                            items_h = [hilite_html(it, rx) for it in items]
                        else:
                            items_h = [it for it in items]
                        excel_val = "\n".join(items)
                        html_val = "<ul class=\"bullets\">" + "".join(f"<li>{it}</li>" for it in items_h) + "</ul>"

            # NaN → tiret cadratin
            if not excel_val or str(excel_val).strip() == '' or str(excel_val).lower() == 'nan':
                excel_val = '—'
            if not html_val or str(html_val).strip() == '' or str(html_val).lower() == 'nan':
                html_val = '—'

            excel_row[label] = excel_val
            # Ajout classe largeur
            cls = COL_WIDTH_CLASS.get(key, 'col-def')
            html_row[label] = f"<div class=\"{cls}\">{html_val}</div>"

        excel_rows.append(excel_row)
        html_rows.append(html_row)

    excel_df = pd.DataFrame(excel_rows, columns=[DISPLAY_LABEL[k] for k in ALL_KEYS_ORDER])
    html_df = pd.DataFrame(html_rows, columns=[DISPLAY_LABEL[k] for k in ALL_KEYS_ORDER])

    # Construction de la table HTML (sans échappement)
    thead = ''.join(f'<th>{DISPLAY_LABEL[k]}</th>' for k in ALL_KEYS_ORDER)
    rows_html = []
    for _, r in html_df.iterrows():
        tds = ''.join(f'<td>{r[c]}</td>' for c in html_df.columns)
        rows_html.append(f'<tr>{tds}</tr>')
    table_html = f"<table class=\"table\"><thead><tr>{thead}</tr></thead><tbody>" + ''.join(rows_html) + "</tbody></table>"
    return html_df, excel_df, table_html

# ──────────────────────────────────────────────────────────────────────────────
#  Export Excel
# ──────────────────────────────────────────────────────────────────────────────

def export_excel(excel_df: pd.DataFrame, article: str) -> str:
    ts = int(time.time())
    path = f"/tmp/filtrage_{ts}.xlsx"
    with pd.ExcelWriter(path, engine='openpyxl') as xw:
        # Insérer une première ligne « Article filtré : ... »
        sheet = 'Filtre'
        excel_df.to_excel(xw, index=False, sheet_name=sheet, startrow=1)
        ws = xw.book[sheet]
        ws.freeze_panes = ws['A3']  # fige l’en‑tête
        ws['A1'] = f"Article filtré : {article}"
        # Largeurs auto
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(excel_df.columns, start=1):
            max_len = max([len(str(col))] + [len(str(v)) for v in excel_df[col].fillna('—').tolist()])
            ws.column_dimensions[get_column_letter(idx)].width = min(70, max(12, max_len + 2))
    return f"/download?path={path}"

# ──────────────────────────────────────────────────────────────────────────────
#  Route principale
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/', methods=['GET','POST'])
def main():
    if request.method == 'GET':
        return render_template_string(HTML, style=STYLE, table_html=None, download_url=None, message=None, ok=True, article='', isolate=False)

    article = (request.form.get('article') or '').strip()
    isolate = bool(request.form.get('isolate'))
    file = request.files.get('file')

    if not file or not article:
        return render_template_string(HTML, style=STYLE, table_html=None, download_url=None, message="Fichier et article requis.", ok=False, article=article, isolate=isolate)

    # Valider extension (openpyxl)
    fname = (file.filename or '').lower()
    if not (fname.endswith('.xlsx') or fname.endswith('.xlsm')):
        return render_template_string(HTML, style=STYLE, table_html=None, download_url=None, message="Format non pris en charge. Fournissez un .xlsx / .xlsm", ok=False, article=article, isolate=isolate)

    try:
        df = read_excel(file.stream)
        colmap = resolve_headers(df)
        # Vérifier disponibilité des 4 colonnes d’intérêt
        missing = [k for k in FILTER_KEYS if k not in colmap]
        if missing:
            return render_template_string(HTML, style=STYLE, table_html=None, download_url=None,
                message=("Colonnes manquantes : " + ", ".join(DISPLAY_LABEL[k] for k in missing)), ok=False, article=article, isolate=isolate)

        html_df, excel_df, table_html = build_tables(df, colmap, article, isolate)
        if excel_df.empty:
            return render_template_string(HTML, style=STYLE, table_html=None, download_url=None, message="Aucune ligne ne contient l’article dans les colonnes cibles.", ok=True, article=article, isolate=isolate)

        download_url = export_excel(excel_df, article)
        # Envelopper table dans un conteneur scrollable
        table_html = f"<div class='table-wrap'>{table_html}</div>"
        return render_template_string(HTML, style=STYLE, table_html=table_html, download_url=download_url, message=None, ok=True, article=article, isolate=isolate)

    except Exception as e:
        return render_template_string(HTML, style=STYLE, table_html=None, download_url=None, message=f"Erreur : {e}", ok=False, article=article, isolate=isolate)


@app.route('/download')
def download():
    path = request.args.get('path')
    if not path or not os.path.exists(path):
        return 'Fichier introuvable.', 404
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
